import openpyxl
import itertools
import re
from collections import OrderedDict

wb = openpyxl.load_workbook('export.xlsx')

links = []

for sheet_name in wb.get_sheet_names():
    ws = wb.get_sheet_by_name(sheet_name)
    rows = ws.max_row
    cols = ws.max_column

    for (r,c) in itertools.product(range(1,rows+1), range(1,cols+1)):
        cell = ws.cell(row=r, column=c)
        value = cell.value
        if(value and isinstance(value, str) and re.match("^=HYPERLINK\(\"(.*?)\"", value)):
            links.append(re.match("^=HYPERLINK\(\"(.*?)\"", value).group(1))


# remove duplicates
links = list(OrderedDict.fromkeys(links))

with open('links.txt', 'w') as f:
    for item in links:
        f.write("%s\n" % item)